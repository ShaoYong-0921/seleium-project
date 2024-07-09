from openpyxl import*
from openpyxl.styles import *
from datetime import datetime

today = datetime.now().strftime("%Y%m%d")
print(f"{today}")
wb = load_workbook(filename=f'{today}.xlsx')
ws = wb.active

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws.column_dimensions[col].width = 15

alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)
style = Font(name='標楷體', size=14, bold=True)

for row in ws.iter_rows():
    for cell in row:
        print(f'{cell = }')
        cell.alignment = alignment
        cell.border = border
        cell.font = style

ws.insert_rows(1)

ws['A1'] = today
date_font_style = Font(name='標楷體', size=14, bold=True)
ws['A1'].font = date_font_style
ws.merge_cells("A1:F1")
ws['A1'].alignment = alignment
ws['A1'].border = border

for sheet in ws:
    print(sheet)

wb.save("test.xlsx")