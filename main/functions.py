from PyPDF2 import PdfReader
from openpyxl.styles import *
from datetime import datetime, timedelta
from openpyxl import load_workbook

import requests

class Functions:
    def __init__(self):
        # self.today = datetime.today().date().strftime("%Y%m%d")
        self.today = (datetime.today().date() - timedelta(days=0)).strftime("%Y%m%d")
        self.status = None

    def getPDF(self):
        # Get today's date
        today = self.today
        # today = datetime.today().date().strftime("%Y%m%d")
        print(f"今天是{today}")

        url = f"https://www.yzu.edu.tw/admin/ga/files/mail/{today}.pdf"

        re = requests.get(url)
        # print(re.status_code)

        if re.status_code == 200:
            with open(f"{today}.pdf", "wb") as file:
                file.write(re.content)
            self.status = True
            # print(f"{status = }")
        else:
            print(f"可是好像還沒有郵件ㄛ code:{re.status_code}")
            self.status = False


    def findData(self):
        with open(f"{self.today}.pdf", "rb") as file:
            dataList = []
            reader = PdfReader(file)
            n = 0
            count = 0 
            for page in reader.pages:
                text = page.extract_text()
                # print(text)
                lines = text.splitlines()
                for line in lines:
                    if line[0] == " ": continue
                    n += 1
                    line = line[len(str(n)):]
                    data = list(map(str, line.split()))
                    if len(data) == 4: data.insert(-1, "")
                    if data[-3] == "教務" :
                        # print(f"{data = }")
                        count += 1
                        data.insert(0, count)
                        dataList.append(data)
                
            print(f"今天有{count}件要送")
            return dataList
    def decoratePDF(self):
        wb = load_workbook(filename=f'{self.today}.xlsx')
        ws = wb.active

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15

        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        style = Font(name='標楷體', size=14, bold=True)

        for row in ws.iter_rows():
            for cell in row:
                # print(f'{cell = }')
                cell.alignment = alignment
                cell.border = border
                cell.font = style

        ws.insert_rows(1)

        ws['A1'] = datetime.today().date().strftime("%Y/%m/%d").replace('/0', '/')
        date_font_style = Font(name='標楷體', size=14, bold=True)
        ws['A1'].font = date_font_style
        ws.merge_cells("A1:F1")
        ws['A1'].alignment = alignment
        for row in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
            ws[row].border =  border


        # for sheet in ws:
        #     print(sheet)

        wb.save("test.xlsx")